"""
Module untuk data processing dan handling Excel files
Termasuk loading, preprocessing, dan saving results
"""

import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import warnings

warnings.filterwarnings('ignore')


class ExcelDataHandler:
    """Kelas untuk handling Excel files"""
    
    def __init__(self, filepath=None):
        """
        Inisialisasi handler
        
        Args:
            filepath: Path ke file Excel
        """
        self.filepath = filepath
        self.data = None
        self.original_data = None
    
    def load_excel(self, filepath=None, sheet_name=0):
        """
        Load data dari file Excel
        
        Args:
            filepath: Path ke file Excel
            sheet_name: Nama sheet atau index (default: 0)
        
        Returns:
            DataFrame
        """
        if filepath:
            self.filepath = filepath
        
        if self.filepath is None:
            raise ValueError("File path must be provided")
        
        if not os.path.exists(self.filepath):
            raise FileNotFoundError(f"File not found: {self.filepath}")
        
        try:
            self.data = pd.read_excel(self.filepath, sheet_name=sheet_name)
            self.original_data = self.data.copy()
            
            print(f"Data loaded successfully!")
            print(f"Shape: {self.data.shape}")
            print(f"\nFirst few rows:")
            print(self.data.head())
            
            return self.data
        
        except Exception as e:
            print(f"Error loading file: {e}")
            return None
    
    def get_info(self):
        """Dapatkan informasi tentang data"""
        if self.data is None:
            raise ValueError("Data not loaded. Use load_excel() first.")
        
        print("\n" + "="*70)
        print("DATA INFORMATION")
        print("="*70)
        print(f"Shape: {self.data.shape}")
        print(f"Columns: {list(self.data.columns)}")
        print(f"\nData Types:\n{self.data.dtypes}")
        print(f"\nMissing Values:\n{self.data.isnull().sum()}")
        print(f"\nBasic Statistics:\n{self.data.describe()}")
    
    def clean_data(self, remove_duplicates=True, remove_na=False):
        """
        Pembersihan data
        
        Args:
            remove_duplicates: Hapus baris duplikat
            remove_na: Hapus baris dengan nilai NA
        """
        if self.data is None:
            raise ValueError("Data not loaded. Use load_excel() first.")
        
        initial_rows = len(self.data)
        
        if remove_duplicates:
            self.data = self.data.drop_duplicates()
            print(f"Removed {initial_rows - len(self.data)} duplicate rows")
        
        if remove_na:
            self.data = self.data.dropna()
            print(f"Removed rows with NA values")
        
        print(f"New shape: {self.data.shape}")
    
    def get_text_column(self, column_name):
        """
        Dapatkan kolom teks dari data
        
        Args:
            column_name: Nama kolom yang berisi teks
        
        Returns:
            List of texts
        """
        if self.data is None:
            raise ValueError("Data not loaded. Use load_excel() first.")
        
        if column_name not in self.data.columns:
            raise ValueError(f"Column '{column_name}' not found in data")
        
        texts = self.data[column_name].astype(str).tolist()
        return texts
    
    def add_column(self, column_name, data):
        """
        Tambah kolom baru ke data
        
        Args:
            column_name: Nama kolom baru
            data: Data untuk kolom (list atau Series)
        """
        if self.data is None:
            raise ValueError("Data not loaded. Use load_excel() first.")
        
        if len(data) != len(self.data):
            raise ValueError(f"Data length mismatch: {len(data)} vs {len(self.data)}")
        
        self.data[column_name] = data
        print(f"Column '{column_name}' added successfully")
    
    def save_to_excel(self, output_path, sheet_name='Sheet1', include_index=False):
        """
        Simpan data ke file Excel
        
        Args:
            output_path: Path untuk file output
            sheet_name: Nama sheet
            include_index: Include index dalam output
        """
        if self.data is None:
            raise ValueError("Data not loaded. Use load_excel() first.")
        
        os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
        
        self.data.to_excel(output_path, sheet_name=sheet_name, index=include_index)
        print(f"Data saved to {output_path}")
    
    def save_with_formatting(self, output_path, sheet_name='Sheet1', include_index=False):
        """
        Simpan data ke Excel dengan formatting
        """
        if self.data is None:
            raise ValueError("Data not loaded. Use load_excel() first.")
        
        os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
        
        self.data.to_excel(output_path, sheet_name=sheet_name, index=include_index)
        
        # Format Excel
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Header formatting
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_path)
        print(f"Formatted data saved to {output_path}")
    
    def export_analysis_results(self, output_path, results_dict):
        """
        Export hasil analisis ke Excel dengan multiple sheets
        
        Args:
            output_path: Path untuk file output
            results_dict: Dictionary dengan {sheet_name: DataFrame}
        """
        os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for sheet_name, df in results_dict.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        print(f"Analysis results exported to {output_path}")


class DataProcessor:
    """Kelas untuk data processing"""
    
    @staticmethod
    def create_sample_dataset(output_path, n_samples=100):
        """
        Buat sample dataset untuk testing
        
        Args:
            output_path: Path untuk file output
            n_samples: Jumlah sample
        """
        sample_texts = [
            "Produk ini sangat bagus dan memuaskan",
            "Kualitas buruk, tidak puas dengan pembelian ini",
            "Harga standar, produk biasa saja",
            "Sangat excellent, highly recommended",
            "Terrible experience, bad service",
            "Produk berkualitas tinggi, worth the price",
            "Saya tidak suka dengan produk ini",
            "Lumayan bagus, sesuai harapan",
            "Absolutely amazing product",
            "Disappointing quality, wasted money"
        ]
        
        labels = [
            'positif', 'negatif', 'netral', 'positif', 'negatif',
            'positif', 'negatif', 'positif', 'positif', 'negatif'
        ]
        
        data = []
        for i in range(n_samples):
            text = sample_texts[i % len(sample_texts)]
            label = labels[i % len(labels)]
            data.append({'id': i+1, 'text': text, 'label': label})
        
        df = pd.DataFrame(data)
        os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
        df.to_excel(output_path, sheet_name='Data', index=False)
        
        print(f"Sample dataset created: {output_path}")
        print(f"Shape: {df.shape}")
        
        return df
    
    @staticmethod
    def split_data(df, text_column, label_column, train_size=0.8):
        """
        Split data menjadi train dan test
        """
        from sklearn.model_selection import train_test_split
        
        texts = df[text_column].tolist()
        labels = df[label_column].tolist()
        
        X_train, X_test, y_train, y_test = train_test_split(
            texts, labels,
            train_size=train_size,
            random_state=42,
            stratify=labels
        )
        
        return X_train, X_test, y_train, y_test


if __name__ == "__main__":
    # Create sample dataset
    sample_path = 'data/sample_data.xlsx'
    DataProcessor.create_sample_dataset(sample_path, n_samples=50)
    
    # Load and process data
    handler = ExcelDataHandler()
    handler.load_excel(sample_path)
    handler.get_info()
    
    # Get texts
    texts = handler.get_text_column('text')
    print(f"\nLoaded {len(texts)} texts")
    print(f"First 3 texts:\n{texts[:3]}")
